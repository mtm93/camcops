#!/usr/bin/env python

"""
camcops_server/cc_modules/cc_tsv.py

===============================================================================

    Copyright (C) 2012-2019 Rudolf Cardinal (rudolf@pobox.com).

    This file is part of CamCOPS.

    CamCOPS is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    CamCOPS is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with CamCOPS. If not, see <http://www.gnu.org/licenses/>.

===============================================================================

**Helper functions/classes for spreadsheet-style tab-separated value (TSV)
exports.**

"""

from collections import OrderedDict
import csv
import io
import logging
from typing import Any, Dict, Iterable, List, Optional, Union
from unittest import TestCase
import zipfile

from cardinal_pythonlib.excel import (
    convert_for_openpyxl,
    excel_to_bytes,
)
from cardinal_pythonlib.logs import BraceStyleAdapter
from odswriter import ODSWriter, Sheet as ODSSheet
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook as XLWorkbook
from openpyxl.worksheet.worksheet import Worksheet as XLWorksheet
# from pyexcel_ods3 import save_data  # poor; use odswriter

log = BraceStyleAdapter(logging.getLogger(__name__))


# =============================================================================
# TSV output holding structures
# =============================================================================

class TsvPage(object):
    """
    Represents a single TSV "spreadsheet".
    """
    def __init__(self, name: str,
                 rows: List[Union[Dict[str, Any], OrderedDict]]) -> None:
        """
        Args:
            name: name for the whole sheet
            rows: list of rows, where each row is a dictionary mapping
                column name to value
        """
        assert name, "Missing name"
        self.name = name
        self.rows = rows
        self.headings = []  # type: List[str]
        for row in rows:
            self._add_headings_if_absent(row.keys())

    def __str__(self) -> str:
        return f"TsvPage: name={self.name}\n{self.get_tsv()}"

    @property
    def empty(self) -> bool:
        """
        Do we have zero rows?
        """
        return len(self.rows) == 0

    def _add_headings_if_absent(self, headings: Iterable[str]) -> None:
        """
        Add any headings we've not yet seen to our list of headings.
        """
        for h in headings:
            if h not in self.headings:
                self.headings.append(h)

    def add_or_set_value(self, heading: str, value: Any) -> None:
        """
        If we contain only a single row, this function will set the value
        for a given column (``heading``) to ``value``.

        Raises:
            :exc:`AssertionError` if we don't have exactly 1 row
        """
        assert len(self.rows) == 1, "add_value can only be used if #rows == 1"
        self._add_headings_if_absent([heading])
        self.rows[0][heading] = value

    def add_or_set_column(self, heading: str, values: List[Any]) -> None:
        """
        Set the column labelled ``heading`` so it contains the values specified
        in ``values``. The length of ``values`` must equal the number of rows
        that we already contain.

        Raises:
            :exc:`AssertionError` if the number of values doesn't match
            the number of existing rows
        """
        assert len(values) == len(self.rows), "#values != #existing rows"
        self._add_headings_if_absent([heading])
        for i, row in enumerate(self.rows):
            row[heading] = values[i]

    def add_or_set_columns_from_page(self, other: "TsvPage") -> None:
        """
        This function presupposes that ``self`` and ``other`` are two pages
        ("spreadsheets") with *matching* rows.

        It updates values or creates columns in ``self`` such that the values
        from all columns in ``other`` are written to the corresponding rows of
        ``self``.

        Raises:
            :exc:`AssertionError` if the two pages (sheets) don't have
            the same number of rows.
        """
        assert len(self.rows) == len(other.rows), "Mismatched #rows"
        self._add_headings_if_absent(other.headings)
        for i, row in enumerate(self.rows):
            for k, v in other.rows[i].items():
                row[k] = v

    def add_rows_from_page(self, other: "TsvPage") -> None:
        """
        Add all rows from ``other`` to ``self``.
        """
        self._add_headings_if_absent(other.headings)
        self.rows.extend(other.rows)

    def sort_headings(self) -> None:
        """
        Sort our headings internally.
        """
        self.headings.sort()

    def get_tsv(self, dialect: str = "excel-tab") -> str:
        """
        Returns the entire page (sheet) as TSV: one header row and then
        lots of data rows.

        For the dialect, see
        https://docs.python.org/3/library/csv.html#csv.excel_tab.

        See also test code in docstring for
        :func:`camcops_server.cc_modules.cc_convert.tsv_from_query`.
        """
        f = io.StringIO()
        writer = csv.writer(f, dialect=dialect)
        writer.writerow(self.headings)
        for row in self.rows:
            writer.writerow([row.get(h) for h in self.headings])
        return f.getvalue()

    def write_to_xlsx_worksheet(self, ws: XLWorksheet) -> None:
        """
        Writes data from this page to an existing XLSX worksheet.
        """
        ws.append(self.headings)
        for row in self.rows:
            ws.append([convert_for_openpyxl(row.get(h))
                       for h in self.headings])

    def write_to_ods_worksheet(self, ws: ODSSheet) -> None:
        """
        Writes data from this page to an existing ODS sheet.
        """
        ws.writerow(self.headings)
        for row in self.rows:
            ws.writerow([row.get(h) for h in self.headings])


class TsvCollection(object):
    """
    A collection of :class:`camcops_server.cc_modules.cc_tsv.TsvPage` pages
    (spreadsheets), like an Excel workbook.
    """
    def __init__(self) -> None:
        self.pages = []  # type: List[TsvPage]

    def __str__(self) -> str:
        return (
            "TsvCollection:\n" +
            "\n\n".join(page.get_tsv() for page in self.pages)
        )

    def page_with_name(self, page_name: str) -> Optional[TsvPage]:
        """
        Returns the page with the specific name, or ``None`` if no such
        page exists.
        """
        return next((page for page in self.pages if page.name == page_name),
                    None)

    def add_page(self, page: TsvPage) -> None:
        """
        Adds a new page to our collection. If the new page has the same name
        as an existing page, rows from the new page are added to the existing
        page. Does nothing if the new page is empty.
        """
        if page.empty:
            return
        existing_page = self.page_with_name(page.name)
        if existing_page:
            # Blend with existing page
            existing_page.add_rows_from_page(page)
        else:
            # New page
            self.pages.append(page)

    def add_pages(self, pages: List[TsvPage]) -> None:
        """
        Adds all ``pages`` to our collection, via :func:`add_page`.
        """
        for page in pages:
            self.add_page(page)

    def sort_headings_within_all_pages(self) -> None:
        """
        Sort headings within each of our pages.
        """
        for page in self.pages:
            page.sort_headings()

    def sort_pages(self) -> None:
        """
        Sort our pages by their page name.
        """
        self.pages.sort(key=lambda p: p.name)

    def get_page_names(self) -> List[str]:
        """
        Return a list of the names of all our pages.
        """
        return [p.name for p in self.pages]

    def get_tsv_file(self, page_name: str) -> str:
        """
        Returns a TSV file for a named page.

        Raises:
            :exc:`AssertionError` if the named page does not exist

        """
        page = self.page_with_name(page_name)
        assert page is not None, f"No such page with name {page_name}"
        return page.get_tsv()

    def as_zip(self, encoding: str = "utf-8") -> bytes:
        """
        Returns the TSV collection as a ZIP file containing TSV files.

        Args:
            encoding: encoding to use when writing the TSV files
        """
        with io.BytesIO() as memfile:
            with zipfile.ZipFile(memfile, "w") as z:
                # Write to ZIP.
                # If there are no valid task instances, there'll be no TSV;
                # that's OK.
                for filename_stem in self.get_page_names():
                    tsv_filename = filename_stem + ".tsv"
                    tsv_contents = self.get_tsv_file(page_name=filename_stem)
                    z.writestr(tsv_filename, tsv_contents.encode(encoding))
            zip_contents = memfile.getvalue()
        return zip_contents

    def as_xlsx(self) -> bytes:
        """
        Returns the TSV collection as an XLSX (Excel) file.
        """
        wb = XLWorkbook()

        # we may have zero pages if there were no rows
        if len(self.pages) > 0:
            wb.remove(wb.active)  # remove the autocreated blank sheet

        for page in self.pages:
            ws = wb.create_sheet(title=self.get_sheet_title(page))
            page.write_to_xlsx_worksheet(ws)

        return excel_to_bytes(wb)

    def get_sheet_title(self, page: TsvPage) -> str:
        title = page.name

        if len(title) > 31:
            title = f"{title[:28]}..."

        return title

    def as_ods(self) -> bytes:
        """
        Returns the TSV collection as an ODS (OpenOffice spreadsheet document)
        file.
        """
        with io.BytesIO() as memfile:
            with ODSWriter(memfile) as odsfile:
                for page in self.pages:
                    sheet = odsfile.new_sheet(name=page.name)
                    page.write_to_ods_worksheet(sheet)
            contents = memfile.getvalue()
        return contents


class TsvCollectionTests(TestCase):
    def test_xlsx_created_from_zero_rows(self) -> None:
        page = TsvPage(name="test", rows=[])
        coll = TsvCollection()
        coll.add_page(page)

        output = coll.as_xlsx()

        # https://en.wikipedia.org/wiki/List_of_file_signatures
        self.assertEqual(output[0], 0x50)
        self.assertEqual(output[1], 0x4B)
        self.assertEqual(output[2], 0x03)
        self.assertEqual(output[3], 0x04)

    def test_xlsx_worksheet_names_are_page_names(self) -> None:
        page1 = TsvPage(name="name 1",
                        rows=[{"test data 1": "row 1"}])
        page2 = TsvPage(name="name 2",
                        rows=[{"test data 2": "row 1"}])
        page3 = TsvPage(name="name 3",
                        rows=[{"test data 3": "row 1"}])
        coll = TsvCollection()

        coll.add_pages([page1, page2, page3])

        data = coll.as_xlsx()
        buffer = io.BytesIO(data)
        wb = load_workbook(buffer)
        self.assertEqual(
            wb.get_sheet_names(),
            [
                "name 1",
                "name 2",
                "name 3",
            ]
        )

    def test_page_name_exactly_31_chars_not_truncated(self) -> None:
        page = TsvPage(name="abcdefghijklmnopqrstuvwxyz78901",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()

        self.assertEqual(
            coll.get_sheet_title(page),
            "abcdefghijklmnopqrstuvwxyz78901"
        )

    def test_page_name_over_31_chars_truncated(self) -> None:
        page = TsvPage(name="abcdefghijklmnopqrstuvwxyz78901234",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()

        self.assertEqual(
            coll.get_sheet_title(page),
            "abcdefghijklmnopqrstuvwxyz78..."
        )
